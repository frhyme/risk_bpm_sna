import pandas as pd
import openpyxl
import networkx as nx
import numpy as np
import difflib
############
def read_excel_first_sheet_and_return_lst(excel_file_name):
    wb=openpyxl.load_workbook(excel_file_name)
    r_sheet=wb.get_sheet_by_name(wb.get_sheet_names()[0])
    row_lst=[]
    for i in range(0, int(r_sheet.max_row)):
        row_lst.append([])
        for j in range(0, int(r_sheet.max_column)):
            temp=r_sheet.cell(row=i+1, column=j+1).value
            temp=str(temp).strip().lower()
            if temp!="none":
                row_lst[i].append(temp)
    return row_lst
    
def return_excel_first_sheet_as_dataframe(excel_name):
    xl=pd.ExcelFile(excel_name)
    return xl.parse(xl.sheet_names[0])
def return_kwds_lst_from_str(kwd_str_lst):
    return_lst=[]#[[a,b,c]]
    for i in range(0, len(kwd_str_lst)):
        return_lst.append(kwd_str_lst[i].split(";"))
        for j in range(0, len(return_lst[i])):
            return_lst[i][j]=return_lst[i][j].strip().lower()
    return return_lst
def return_kwds_lst_from_df(df):
    kwd_str_lst=df.dropna().tolist()
    kwds_lst=return_kwds_lst_from_str(kwd_str_lst)
    return kwds_lst
def return_kwd_relation_lst_from_df(df):
    kwd_str_lst=df.dropna().tolist()
    kwds_lst=return_kwds_lst_from_str(kwd_str_lst)
    return_lst=[]
    for i in range(0, len(kwds_lst)):
        for j1 in range(0, len(kwds_lst[i])-1):
            for j2 in range(j1+1, len(kwds_lst[i])):
                if kwds_lst[i][j1]!=kwds_lst[i][j2]:
                    return_lst.append(sorted([kwds_lst[i][j1], kwds_lst[i][j2]]))
    return return_lst
def return_kwd_relation_lst_from_kwds_lst(kwds_lst):
    return_lst=[]
    for i in range(0, len(kwds_lst)):
        for j1 in range(0, len(kwds_lst[i])-1):
            for j2 in range(j1+1, len(kwds_lst[i])):
                return_lst.append(sorted([kwds_lst[i][j1], kwds_lst[i][j2]]))
    return return_lst
def return_kwd_lst_from_df(df):
    kwd_str_lst=df.dropna().tolist()
    kwds_lst=return_kwds_lst_from_str(kwd_str_lst)
    return_lst=[]
    for i in range(0, len(kwds_lst)):
        for j in range(0, len(kwds_lst[i])):
            return_lst.append(kwds_lst[i][j])
    return return_lst
def return_kwd_lst_from_kwds_lst(kwds_lst):
    return_lst=[]
    for i in range(0, len(kwds_lst)):
        for j in range(0, len(kwds_lst[i])):
            return_lst.append(kwds_lst[i][j])
    return return_lst
def return_kwd_count_from_kwd_lst(kwd_lst):
    sorted_kwd_lst=sorted(kwd_lst)
    key_lst=[]
    count_lst=[]
    
    for i in range(0, len(sorted_kwd_lst)):
        if sorted_kwd_lst[i] in key_lst:
            count_lst[key_lst.index(sorted_kwd_lst[i])]+=1
        else:
            key_lst.append(sorted_kwd_lst[i])
            count_lst.append(1)
    return_lst=[ [key_lst[i], count_lst[i]] for i in range(0, len(key_lst))]
    return_lst.sort(key=lambda tup:tup[1], reverse=True)
    return return_lst
def return_kwd_count_from_kwds_lst(kwds_lst):
    return return_kwd_count_from_kwd_lst( return_kwd_lst_from_kwds_lst(kwds_lst) )
def return_kwd_count_from_df(df):
    return return_kwd_count_from_kwd_lst(return_kwd_lst_from_df(df))
def wrt_two_dim_lst_in_excel(excel_file_name, sheet_name, two_dim_lst):
    wb=openpyxl.load_workbook(filename=excel_file_name)
    w_sheet=wb.create_sheet(title=sheet_name)
    for i in range(0, len(two_dim_lst)):
        for j in range(0, len(two_dim_lst[i])):
            w_sheet.cell(row=i+1, column=j+1, value=two_dim_lst[i][j])
    wb.save(excel_file_name)
def return_kwds_lst_applied_by_rule_lst_from_excel(excel_file_name, kwds_lst):
    rule_lst=read_excel_first_sheet_and_return_lst(excel_file_name)
    for one_rule in rule_lst:
        if one_rule==[]:
            continue
        for i in range(0, len(kwds_lst)):
            for j in range(0, len(kwds_lst[i])):
                if one_rule[0]==kwds_lst[i][j]:
                    kwds_lst[i][j]=one_rule[1]
    return kwds_lst
def write_df_in_excel(excel_name, excel_sheet_name, df):
    writer=pd.ExcelWriter(excel_name, engine="xlsxwriter")
    df.to_excel(writer, sheet_name=excel_sheet_name)
    writer.save()
#################
######main
#kwds_str_lst
#kwds_lst
#kwd_relation_lst: keyword_pair
#kwd_count
###TRIP
def return_str_similarity(str1, str2):
    return difflib.SequenceMatcher(a=str1, b=str2).ratio()
def make_keyword_change_rule_based_on_similarity(input_excel_name, output_excel_name):
    #similarity를 기준으로 비슷한 것에 대해서는 sort하여, rule로 만들어줌
    df=return_excel_first_sheet_as_dataframe(input_excel_name)
    ##dataframe to dictionary
    yearly_kwds_dict={}
    for i in range(int(df["Year"].min()), int(df["Year"].max())+1):
        target_df=df[df["Year"]==i]["Author Keywords"]
        yearly_kwds_dict[i]=return_kwds_lst_from_df( target_df )

    start_year=2006
    end_year=2015
    
    total_kwd_lst=[ yearly_kwds_dict[year][j][k] for year in range(start_year, end_year+1) for j in range(0, len(yearly_kwds_dict[year])) for k in range(0, len(yearly_kwds_dict[year][j]))]
    total_kwd_count_lst = return_kwd_count_from_kwd_lst(total_kwd_lst)

    kwd_dict_key=[ total_kwd_count_lst[i][0] for i in range(0, len(total_kwd_count_lst)) ]
    kwd_dict_value=[ total_kwd_count_lst[i][1] for i in range(0, len(total_kwd_count_lst)) ]

    kwd_count_dict={kwd_dict_key[i]: kwd_dict_value[i] for i in range(0, len(kwd_dict_key))}
    
    unique_kwd_lst=sorted([elem[0] for elem in total_kwd_count_lst])
    print(len(unique_kwd_lst))

    """
    kwd_sim_lst=[]
    for i in range(0, len(unique_kwd_lst)):
        for j in range(i+1, len(unique_kwd_lst)):
            temp=return_str_similarity( unique_kwd_lst[i], unique_kwd_lst[j] ) 
            if temp > 0.5:
                kwd_sim_lst.append( temp )
    print(np.mean(kwd_sim_lst))
    print(np.std(kwd_sim_lst))
    kwd_sim_threshold=np.mean(kwd_sim_lst)+2*np.std(kwd_sim_lst)
    print(kwd_sim_threshold)
    """
    kwd_sim_threshold=0.9
    
    from_kwd=[]
    from_kwd_count=[]
    to_kwd=[]
    to_kwd_count=[]
            
    #threshold 이상의 키워드 rule 도출
    for i in range(0, len(unique_kwd_lst)):
        for j in range(i+1, len(unique_kwd_lst)):
            sim_value=return_str_similarity( unique_kwd_lst[i], unique_kwd_lst[j] )
            if sim_value==1:
                continue
            elif sim_value>=kwd_sim_threshold:
                if kwd_count_dict[unique_kwd_lst[i]] <  kwd_count_dict[unique_kwd_lst[j]]:
                    from_kwd.append( unique_kwd_lst[i] )
                    from_kwd_count.append( kwd_count_dict[unique_kwd_lst[i]] )
                    to_kwd.append( unique_kwd_lst[j] )
                    to_kwd_count.append( kwd_count_dict[unique_kwd_lst[j]] )
                elif kwd_count_dict[unique_kwd_lst[i]] ==  kwd_count_dict[unique_kwd_lst[j]]:
                    if kwd_count_dict[unique_kwd_lst[i]] > kwd_count_dict[unique_kwd_lst[j]]:
                        from_kwd.append( unique_kwd_lst[i] )
                        from_kwd_count.append( kwd_count_dict[unique_kwd_lst[i]] )
                        to_kwd.append( unique_kwd_lst[j] )
                        to_kwd_count.append( kwd_count_dict[unique_kwd_lst[j]] )
                    else:
                        from_kwd.append( unique_kwd_lst[j] )
                        from_kwd_count.append( kwd_count_dict[unique_kwd_lst[j]] )
                        to_kwd.append( unique_kwd_lst[i] )
                        to_kwd_count.append( kwd_count_dict[unique_kwd_lst[i]] )                        
                else:
                    from_kwd.append( unique_kwd_lst[j] )
                    from_kwd_count.append( kwd_count_dict[unique_kwd_lst[j]] )
                    to_kwd.append( unique_kwd_lst[i] )
                    to_kwd_count.append( kwd_count_dict[unique_kwd_lst[i]] )
            else:
                break

    df=pd.DataFrame( {"1_from":from_kwd, "2_to":to_kwd, "3_from_fre":from_kwd_count, "4_to_fre":to_kwd_count }, index=None)
    #df=pd.DataFrame( {"from":from_kwd, "to":to_kwd }, index=None)

    writer=pd.ExcelWriter(output_excel_name, engine="xlsxwriter")
    excel_sheet_name="rule_lst"
    df.to_excel(writer, sheet_name=excel_sheet_name, header=False, index=False)
    writer.save()
    print("make rule complete")
def make_keyword_report(input_excel_name, output_excel_name, kwd_change_rule_excel_name, kwd_remove_rule_excel_name):
    ####input_excel_name, output_excel_name, rule_lst 이 셋만 파일명으로 받으면, trip, twip 이 두가지를 동시에 진행할 수 있음.
    ###basic information: summary or describe could be used???
    ##data 처리를 초반에 다하고, 그 뒤에 마무리하는 식으로 진행하는게 좋음. 지금은 매번 rule_lst를 적용하는 식으로 진행해서, 문제가 많음
    #1) input_excel_name
    #2) kwd_change_rule_lst
    #3) kwd_remove_rule_lst
    ##main_kwd_set, remove_kwd_set
    kwd_change_rule_lst=read_excel_first_sheet_and_return_lst(kwd_change_rule_excel_name)
    #kwd_remove_rule_lst=kwd_remove_rule_excel_name
    kwd_remove_rule_lst=["trip", "trip steel", "twip", "steel", "twip steel", "steels"]
    #kwd_remove_rule_lst=[]
    ####remove 부분 처리해줘야 함

    start_year=2006
    end_year=2015

    df=return_excel_first_sheet_as_dataframe(input_excel_name)

    ##dataframe to dictionary
    yearly_kwds_dict={}
    yearly_arc_dict={}
    for i in range(int(df["Year"].min()), int(df["Year"].max())+1):
        #print(i)
        target_df=df[df["Year"]==i]["Author Keywords"]
        yearly_kwds_dict[i]=return_kwds_lst_from_df( target_df )
        yearly_arc_dict[i]=return_kwd_relation_lst_from_df( target_df )
    k=0
    for year in range(start_year, end_year):
        for i in range(0, len(yearly_kwds_dict[year])):
            k=k+len(yearly_kwds_dict[year][i])
    print(k)
    
    #keyword change in yearly_arc_dict
    for one_rule in kwd_change_rule_lst:
        if one_rule==[]:
            continue
        for year in range(int(df["Year"].min()), int(df["Year"].max())+1):
            for i in range(0, len(yearly_arc_dict[year])):
                for j in range(0, len(yearly_arc_dict[year][i])):
                    if yearly_arc_dict[year][i][j]==one_rule[0]:
                        yearly_arc_dict[year][i][j]=one_rule[1]
            for i in range(0, len(yearly_kwds_dict[year])):
                for j in range(0, len(yearly_kwds_dict[year][i])):
                    if yearly_kwds_dict[year][i][j]==one_rule[0]:
                        yearly_kwds_dict[year][i][j]=one_rule[1]

    #keyword remove in yearly_arc_dict and yearly_kwds_dict
    
    for remove_target in kwd_remove_rule_lst:
        for year in range(start_year, end_year+1):
            return_lst=[]
            for arc in yearly_arc_dict[year]:
                if remove_target not in arc:
                    return_lst.append(arc)
            yearly_arc_dict[year]=return_lst

            for kwds_lst in yearly_kwds_dict[year]:
                for kwd in kwds_lst:
                    if remove_target ==kwd:
                        kwds_lst.remove(kwd)
    #####
    total_kwd_relation_lst=[]
    for year in range(start_year, end_year+1):
        total_kwd_relation_lst = total_kwd_relation_lst + yearly_arc_dict[year]
    excel_sheet_name_total_kwd_relation = "기간 내 arc list"
    df_total_kwd_relation = pd.DataFrame({ "source":[ item[0] for item in total_kwd_relation_lst ], "target":[ item[1] for item in total_kwd_relation_lst ] })
    #0) 기간내  키워드 빈도수 표 
    total_kwd_lst=[ yearly_kwds_dict[year][j][k] for year in range(start_year, end_year+1) for j in range(0, len(yearly_kwds_dict[year])) for k in range(0, len(yearly_kwds_dict[year][j]))]
    total_kwd_count_lst = return_kwd_count_from_kwd_lst(total_kwd_lst)

    df_kwd_index=[ total_kwd_count_lst[i][0] for i in range(0, len(total_kwd_count_lst)) ]
    df_frequency_column=[ total_kwd_count_lst[i][1] for i in range(0, len(total_kwd_count_lst)) ]
    excel_sheet_name0="키워드 빈도수 표"
    df0=pd.DataFrame({"frequency": df_frequency_column}, index=df_kwd_index)
    print("0) 기간 내 키워드 빈도수 표 complete")
    
    #1) 연도별 빈도수 상위 50개 키워드  변화 표
    excel_sheet_name1="연도별 빈도수 상위 50개 키워드"
    temp_dict={}
    
    for year in range(start_year, end_year+1):
        #print(year)
        kwd_count_lst = return_kwd_count_from_kwds_lst(yearly_kwds_dict[year])
        yearly_kwd_lst = []
        until_nth_rank=50
        #until_nth_rank = 50 if 50<len(kwd_count_lst) else len(kwd_count_lst)
        for i in range(0, until_nth_rank):
            if i>=len(kwd_count_lst):
                yearly_kwd_lst.append("")
            else:
                yearly_kwd_lst.append(kwd_count_lst[i][0])
        temp_dict[year]=yearly_kwd_lst
    df1=pd.DataFrame(temp_dict)
    print("1) 전체 기간 빈도수 상위 50개 키워드를 대상으로 연도별 빈도수 변화 표 complete")

    #2) 전체 기간 상위 50개 키워드 빈도수 변화
    excel_sheet_name2="기간내 상위 50개 키워드 연도별 빈도수 변화"
    temp_dict={}
    
    kwd_count_50_lst = [ total_kwd_count_lst[i][0] for i in range(0, 50)]

    for year in range(start_year, end_year+1):
        count_lst=[ 0 for i in range(0, 50) ]
        for i in range(0, len(kwd_count_50_lst)):
            for j in range(0, len( yearly_kwds_dict[year] ) ):
                for k in range(0, len(yearly_kwds_dict[year][j])):
                    #print(year, i, j, k)
                    if kwd_count_50_lst[i]==yearly_kwds_dict[year][j][k]:
                        count_lst[i]=count_lst[i]+1
                        #print(year, kwd_count_50_lst[i], yearly_arc_dict[year][j][k], count_num)
        temp_dict[year]=count_lst
    df2=pd.DataFrame(temp_dict, index=kwd_count_50_lst)
    print("2) 전체 기간 상위 50개 키워드 빈도수 변화 complete")

    #3) 연도별 deg/close/betweeness centrality 상위 키워드변화"
    excel_sheet_name3="Deg_cent_kwd_yearly"
    excel_sheet_name4="Close_cent_kwd_yearly"
    excel_sheet_name5="Between_cent_kwd_yearly"

    deg_dict = {}
    close_dict = {}
    betw_dict = {}

    for year in range(start_year, end_year+1):
        multi_g = nx.MultiGraph()
        multi_g.add_edges_from(yearly_arc_dict[year])
        kwd_deg_cen_dict = nx.degree_centrality(multi_g)
        kwd_close_cen_dict = nx.closeness_centrality(multi_g)
        kwd_betw_cen_dict = nx.betweenness_centrality(multi_g)
        
        kwd_deg_cen_lst = sorted( [ [key, kwd_deg_cen_dict[key] ] for key in kwd_deg_cen_dict.keys() ], key=lambda x: x[1], reverse=True)
        kwd_close_cen_lst = sorted( [ [key, kwd_close_cen_dict[key] ] for key in kwd_close_cen_dict.keys() ], key=lambda x: x[1], reverse=True)
        kwd_betw_cen_lst = sorted( [ [key, kwd_betw_cen_dict[key] ] for key in kwd_betw_cen_dict.keys() ], key=lambda x: x[1], reverse=True)

        deg_dict[year] = [ kwd_deg_cen_lst[i][0] if len(kwd_deg_cen_lst) > i else "" for i in range(0, 50) ]
        close_dict[year] = [ kwd_close_cen_lst[i][0] if len(kwd_close_cen_lst) > i else ""  for i in range(0, 50) ]
        betw_dict[year] = [ kwd_betw_cen_lst[i][0] if len(kwd_betw_cen_lst) > i else ""  for i in range(0, 50) ]

    df3=pd.DataFrame(deg_dict)
    df4=pd.DataFrame(close_dict)
    df5=pd.DataFrame(betw_dict)
    print("3) 연도별 deg/close/betweeness centrality 상위 키워드변화 complete")

    #4) 기간내 상위 50개 키워드, centrality 연도별 변화 표
    multi_g=nx.MultiGraph()
    for year in range(start_year, end_year+1):
        multi_g.add_edges_from(yearly_arc_dict[year])
    kwd_deg_cen_dict = nx.degree_centrality(multi_g)
    kwd_close_cen_dict = nx.closeness_centrality(multi_g)
    kwd_betw_cen_dict = nx.betweenness_centrality(multi_g)

temp=[]
for i in range(0, 10):
    temp.append(i)
    
    kwd_deg_cen_lst = sorted( [ [key, kwd_deg_cen_dict[key] ] for key in kwd_deg_cen_dict.keys() ], key=lambda x: x[], reverse=True)
    kwd_close_cen_lst = sorted( [ [key, kwd_close_cen_dict[key] ] for key in kwd_close_cen_dict.keys() ], key=lambda x: x[1], reverse=True)
    kwd_betw_cen_lst = sorted( [ [key, kwd_betw_cen_dict[key] ] for key in kwd_betw_cen_dict.keys() ], key=lambda x: x[1], reverse=True)

    deg_lst = [ kwd_deg_cen_lst[i][0] for i in range(0, 50) ]
    close_lst = [ kwd_close_cen_lst[i][0] for i in range(0, 50) ]
    betw_lst = [ kwd_betw_cen_lst[i][0] for i in range(0, 50) ]

    excel_sheet_name6="상위 50키워드 yearly deg.cen 변화"
    excel_sheet_name7="상위 50키워드 yearly close.cen 변화"
    excel_sheet_name8="상위 50키워드 yearly between.cen 변화"
    print("4) 전체 키워드 centrality(degree/closeness/betweenness) 상위 50키워드 도출 complete")

    temp_dict1={}
    temp_dict2={}
    temp_dict3={}
    for year in range(start_year, end_year+1):
        deg_value_lst=[]
        close_value_lst=[]
        betw_value_lst=[]

        multi_g=nx.MultiGraph()
        multi_g.add_edges_from(yearly_arc_dict[year])

        kwd_deg_cen_dict = nx.degree_centrality(multi_g)
        kwd_close_cen_dict = nx.closeness_centrality(multi_g)
        kwd_betw_cen_dict = nx.betweenness_centrality(multi_g)

        for i in range(0, 50):
            if deg_lst[i] in kwd_deg_cen_dict.keys():
                deg_value_lst.append( kwd_deg_cen_dict[ deg_lst[i] ] )
            else:
                deg_value_lst.append( 0 )
                
            if close_lst[i] in kwd_close_cen_dict.keys():
                close_value_lst.append( kwd_close_cen_dict[ close_lst[i] ] )
            else:
                close_value_lst.append( 0 )
                
            if betw_lst[i] in kwd_betw_cen_dict.keys():
                betw_value_lst.append( kwd_betw_cen_dict[ betw_lst[i] ] )
            else:
                betw_value_lst.append( 0 )
        temp_dict1[year]=deg_value_lst
        temp_dict2[year]=close_value_lst
        temp_dict3[year]=betw_value_lst
        
    df6=pd.DataFrame(temp_dict1, index=deg_lst)
    df7=pd.DataFrame(temp_dict2, index=close_lst)
    df8=pd.DataFrame(temp_dict3, index=betw_lst)

    ###report complete
    #write yearly arc lists in excel 
    writer=pd.ExcelWriter(output_excel_name, engine="xlsxwriter")
    for year in range(start_year, end_year+1):
        df=pd.DataFrame(yearly_arc_dict[year], columns=["source", "target"])
        df.to_excel(writer, sheet_name=str(year))
    df_total_kwd_relation.to_excel(writer, sheet_name=excel_sheet_name_total_kwd_relation, index=False, header=False)
    
    df0.to_excel(writer, sheet_name=excel_sheet_name0)
    df1.to_excel(writer, sheet_name=excel_sheet_name1)
    df2.to_excel(writer, sheet_name=excel_sheet_name2)
    df3.to_excel(writer, sheet_name=excel_sheet_name3)
    df4.to_excel(writer, sheet_name=excel_sheet_name4)
    df5.to_excel(writer, sheet_name=excel_sheet_name5)
    df6.to_excel(writer, sheet_name=excel_sheet_name6)
    df7.to_excel(writer, sheet_name=excel_sheet_name7)
    df8.to_excel(writer, sheet_name=excel_sheet_name8)

    writer.save()

    print("All complete")
#####################################
###############function definition complete        
trip_excel_name="TRIP_STEEL(selected_columns).xlsx"
trip_output_name="TRIP_keyword_report.xlsx"
trip_change_rule_excel_name="rule_lst.xlsx"
trip_remove_rule_excel_name=""
#make_keyword_change_rule_based_on_similarity(trip_excel_name, "trip_rule_lst.xlsx")
make_keyword_report(trip_excel_name, trip_output_name, "rule_lst_v2.xlsx", trip_remove_rule_excel_name)

twip_excel_name="TWIP_STEEL(selected_columns).xlsx"
twip_output_name="TWIP_keyword_report.xlsx"
twip_change_rule_excel_name="twip_rule_lst.xlsx"
twip_remove_rule_excel_name=""
#make_keyword_change_rule_based_on_similarity(twip_excel_name, "twip_rule_lst.xlsx")
make_keyword_report(twip_excel_name, twip_output_name, "rule_lst_v2.xlsx", twip_remove_rule_excel_name)

"""
###TWIP
#twip_excel_name="TWIP_STEEL(selected_columns).xlsx"
#twip_df=return_excel_first_sheet_as_dataframe(twip_excel_name)
#print(twip_df.columns)


#print(trip_df[trip_df["Year"]==2009]["Author Keywords"])
#dropna => delete row having nan, tolist==> Series to list




"""




